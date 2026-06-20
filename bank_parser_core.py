"""
bank_parser_core.py
Abstract base class for all bank statement parsers.

Every bank-specific parser subclasses BankStatementParser and only needs
to implement parse_line() — the logic for turning one OCR'd text line
into a transaction dict. Everything else (OCR, Excel writing, balance
validation) is shared.
"""

import re
import sys
from abc import ABC, abstractmethod
from pathlib import Path

import pytesseract
from pdf2image import convert_from_path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


# ── Shared regex building blocks (banks can reuse or override) ────────────────

DR_CR_RE = re.compile(r'\b(Cr|Dr|CR|DR)\b')

CONF_THRESHOLD = 60   # OCR confidence below this flags the page


class BankStatementParser(ABC):
    """
    Abstract base for a single bank's statement parser.

    Subclasses MUST implement:
        - BANK_NAME            class attribute, e.g. "Federal Bank"
        - DETECT_KEYWORDS       list of strings that identify this bank in OCR text
        - parse_line(line)      turn one OCR text line into a transaction dict or None

    Subclasses MAY override:
        - SKIP_LINE_MARKERS     strings that mean "this line is header/footer noise"
        - extract_metadata(text) pull account holder / number / IFSC etc from page 1
    """

    BANK_NAME = "Unknown Bank"
    DETECT_KEYWORDS = []   # e.g. ["Federal Bank", "federalbank.co.in"]

    SKIP_LINE_MARKERS = [
        'Statement of Account', 'Opening Balance', 'Date Value Date',
        'Tran Type', 'Page ', 'Cheque', 'Withdrawals', 'Deposits',
        'Balance', 'DR /CR', 'DR/CR', 'Corporate Office',
    ]

    def __init__(self, tesseract_cmd=None, poppler_path=None, dpi=200):
        self.tesseract_cmd = tesseract_cmd
        self.poppler_path  = poppler_path
        self.dpi            = dpi
        if tesseract_cmd:
            pytesseract.pytesseract.tesseract_cmd = tesseract_cmd

    # ── Abstract method every bank must implement ──────────────────────────

    @abstractmethod
    def parse_line(self, line):
        """
        Parse a single OCR text line into a transaction dict, or None
        if the line doesn't contain a transaction.

        Expected dict keys:
            date, value_date, particulars, tran_type, tran_id,
            withdrawals, deposits, balance, dr_cr
        """
        raise NotImplementedError

    # ── Optional override ──────────────────────────────────────────────────

    def extract_metadata(self, text):
        """Pull account holder, account number etc from page 1 text. Override per bank."""
        meta = {}
        patterns = {
            'name':            r'Name\s*:\s*([^\n]+)',
            'account_number':  r'Account Number\s*:\s*(\S+)',
            'period':          r'Statement of Account for the period\s+(\S+\s+to\s+\S+)',
            'ifsc':            r'IFSC\s*:\s*(\S+)',
        }
        for key, pat in patterns.items():
            m = re.search(pat, text, re.IGNORECASE)
            if m:
                meta[key] = m.group(1).strip()
        return meta

    # ── Shared OCR ──────────────────────────────────────────────────────────

    def ocr_page(self, img):
        """Return (text, avg_conf) for a PIL image using PSM 4."""
        data = pytesseract.image_to_data(
            img, config='--psm 4 --oem 3', output_type=pytesseract.Output.DICT
        )
        confs = [c for c in data['conf'] if c != -1]
        avg_conf = sum(confs) / len(confs) if confs else 0
        text = pytesseract.image_to_string(img, config='--psm 4 --oem 3')
        return text, avg_conf

    # ── Shared multi-line transaction assembly ─────────────────────────────

    def parse_page_text(self, text, page_avg_conf):
        """
        Parse raw OCR text from one page into a list of transaction dicts.
        Handles multi-line particulars by accumulating continuation lines.
        Calls self.parse_line() for each candidate line — bank-specific.
        """
        transactions = []
        current = None

        for raw_line in text.splitlines():
            line = raw_line.strip()
            if not line:
                continue

            if any(marker in line for marker in self.SKIP_LINE_MARKERS):
                if current:
                    transactions.append(current)
                    current = None
                continue

            parsed = self.parse_line(line)

            if parsed and parsed.get('date'):
                if current:
                    transactions.append(current)
                parsed['_low_conf'] = page_avg_conf < CONF_THRESHOLD
                current = parsed
            else:
                # Continuation line — append to current particulars
                if current and re.search(r'[a-zA-Z@/]', line):
                    current['particulars'] = (current['particulars'] + ' ' + line).strip()

        if current:
            transactions.append(current)

        return transactions

    # ── Shared balance-chain validation ─────────────────────────────────────

    def validate_balances(self, transactions):
        """
        Walk the transaction list and flag rows where
        prev_balance ± amount != current_balance.
        Adds a '_balance_flag' key to each transaction.
        """
        prev_balance = None
        for tx in transactions:
            bal = tx.get('balance')
            wd  = tx.get('withdrawals')
            dep = tx.get('deposits')

            if bal is None:
                tx['_balance_flag'] = 'missing balance'
                continue

            if prev_balance is None:
                tx['_balance_flag'] = None
                prev_balance = bal
                continue

            if wd is not None:
                expected = round(prev_balance - wd, 2)
            elif dep is not None:
                expected = round(prev_balance + dep, 2)
            else:
                tx['_balance_flag'] = 'missing amount'
                prev_balance = bal
                continue

            if abs(expected - bal) <= 0.02:
                tx['_balance_flag'] = None
            else:
                # The amount was likely put in the wrong bucket (withdrawal
                # vs deposit). Before giving up, try the OTHER interpretation:
                # if swapping wd <-> dep makes the balance reconcile, the
                # OCR's punctuation-based guess was wrong — fix it.
                swapped_expected = (
                    round(prev_balance + wd, 2) if wd is not None
                    else round(prev_balance - dep, 2)
                )
                if abs(swapped_expected - bal) <= 0.02:
                    amount = wd if wd is not None else dep
                    if wd is not None:
                        tx['withdrawals'] = None
                        tx['deposits']    = amount
                    else:
                        tx['deposits']    = None
                        tx['withdrawals'] = amount
                    tx['_balance_flag'] = 'auto-corrected: wrong column, fixed via balance check'
                else:
                    tx['_balance_flag'] = f'expected {expected}, got {bal}'

            prev_balance = bal

        return transactions

    # ── Main entry point ────────────────────────────────────────────────────

    def parse_pdf(self, pdf_path, progress_callback=None):
        """
        Run the full pipeline: convert PDF → OCR each page → parse →
        validate balances. Returns (transactions, metadata).

        progress_callback(current_page, total_pages) is called after each
        page if provided — useful for a GUI progress bar.
        """
        pdf_path = Path(pdf_path)
        if not pdf_path.exists():
            raise FileNotFoundError(f"PDF not found: {pdf_path}")

        kwargs = {'dpi': self.dpi}
        if self.poppler_path:
            kwargs['poppler_path'] = self.poppler_path

        images = convert_from_path(str(pdf_path), **kwargs)
        total  = len(images)

        all_transactions = []
        meta           = {}
        meta_extracted = False
        low_conf_pages = []

        for i, img in enumerate(images, 1):
            text, avg_conf = self.ocr_page(img)

            if not meta_extracted:
                m = self.extract_metadata(text)
                if m:
                    meta = m
                    meta_extracted = True

            if avg_conf < CONF_THRESHOLD:
                low_conf_pages.append(i)

            page_txns = self.parse_page_text(text, avg_conf)
            all_transactions.extend(page_txns)

            if progress_callback:
                progress_callback(i, total)

        all_transactions = self.validate_balances(all_transactions)

        # closing balance = last transaction's balance
        for tx in reversed(all_transactions):
            if tx.get('balance') is not None:
                meta['closing_balance'] = tx['balance']
                break

        meta['bank']           = self.BANK_NAME
        meta['low_conf_pages'] = low_conf_pages
        meta['total_pages']    = total

        return all_transactions, meta

    # ── Excel column definition — OVERRIDE these in each bank subclass ──────
    #
    # EXCEL_HEADERS: column titles in the order you want them to appear,
    #     matching that bank's own statement layout.
    # EXCEL_COL_WIDTHS: width per column letter (A, B, C, ...). Must have
    #     one entry per header, in the same order.
    # EXCEL_NUMBER_COLS: 1-indexed column numbers that should get the
    #     #,##0.00 number format (your amount/balance columns).
    #
    # to_excel_row(tx) maps one transaction dict to a list of values in
    # EXCEL_HEADERS order. Override it alongside EXCEL_HEADERS — they must
    # stay in sync. The default below matches the generic field names used
    # by parse_line() in this base class; if your subclass uses different
    # dict keys, override to_excel_row() to match.

    EXCEL_HEADERS = ['Date', 'Value Date', 'Particulars', 'Tran Type', 'Tran ID',
                      'Withdrawals', 'Deposits', 'Balance', 'DR/CR']

    EXCEL_COL_WIDTHS = {'A': 14, 'B': 14, 'C': 55, 'D': 10, 'E': 16,
                         'F': 14, 'G': 14, 'H': 16, 'I': 8}

    EXCEL_NUMBER_COLS = {6, 7, 8}   # Withdrawals, Deposits, Balance

    def to_excel_row(self, tx):
        """Map a transaction dict to a list of values matching EXCEL_HEADERS order."""
        return [
            tx.get('date', ''), tx.get('value_date', ''), tx.get('particulars', ''),
            tx.get('tran_type', ''), tx.get('tran_id', ''),
            tx.get('withdrawals'), tx.get('deposits'), tx.get('balance'),
            tx.get('dr_cr', ''),
        ]

    def excel_summary_rows(self, transactions, meta):
        """
        Build the (label, value) rows for the Summary sheet.
        Override per bank if you want different totals shown.
        Default sums 'withdrawals' and 'deposits' fields.
        """
        total_wd     = sum(t.get('withdrawals') or 0 for t in transactions)
        total_dep    = sum(t.get('deposits')    or 0 for t in transactions)
        auto_fixed   = sum(1 for t in transactions
                           if (t.get('_balance_flag') or '').startswith('auto-corrected'))
        still_flagged = sum(1 for t in transactions
                            if (t.get('_balance_flag') and not t['_balance_flag'].startswith('auto-corrected'))
                            or t.get('_low_conf'))

        return [
            ('Bank',               meta.get('bank', '')),
            ('Account Holder',     meta.get('name', '')),
            ('Account Number',     meta.get('account_number', '')),
            ('Statement Period',   meta.get('period', '')),
            ('', ''),
            ('Total Transactions', len(transactions)),
            ('Total Withdrawals',  total_wd),
            ('Total Deposits',     total_dep),
            ('Closing Balance',    meta.get('closing_balance', '')),
            ('', ''),
            ('Auto-corrected Rows', auto_fixed),
            ('Flagged Rows',        still_flagged),
            ('Low Confidence Pages', str(meta.get('low_conf_pages', []))),
        ]

    # ── Excel writer (generic — reads structure from the subclass) ──────────

    def write_excel(self, transactions, meta, output_path):
        wb = Workbook()
        ws = wb.active
        ws.title = 'Transactions'
        ws.freeze_panes = 'A2'

        header_fill = PatternFill('solid', fgColor='1F4E79')
        alt_fill    = PatternFill('solid', fgColor='EBF3FB')
        warn_fill   = PatternFill('solid', fgColor='FFF2CC')
        fix_fill    = PatternFill('solid', fgColor='E2EFDA')   # green — auto-corrected
        header_font = Font(name='Arial', bold=True, color='FFFFFF', size=10)
        body_font   = Font(name='Arial', size=9)
        center      = Alignment(horizontal='center', vertical='center')
        left        = Alignment(horizontal='left', vertical='center', wrap_text=True)
        right       = Alignment(horizontal='right', vertical='center')
        thin        = Side(style='thin', color='D0D0D0')
        border      = Border(left=thin, right=thin, top=thin, bottom=thin)

        headers    = self.EXCEL_HEADERS + ['Flag']
        num_cols   = self.EXCEL_NUMBER_COLS
        n_cols     = len(headers)

        for col, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=col, value=h)
            c.font, c.fill, c.alignment, c.border = header_font, header_fill, center, border
        ws.row_dimensions[1].height = 22

        # Build alignment per column: right-align number columns, left-align
        # the widest text-ish column (assume it's the longest header — usually
        # "Particulars"/"Description"), center everything else.
        text_col_idx = max(range(1, len(self.EXCEL_HEADERS) + 1),
                            key=lambda i: len(self.EXCEL_HEADERS[i - 1]))
        aligns = []
        for i in range(1, n_cols + 1):
            if i in num_cols:
                aligns.append(right)
            elif i == text_col_idx:
                aligns.append(left)
            elif i == n_cols:  # Flag column
                aligns.append(left)
            else:
                aligns.append(center)

        for row_idx, tx in enumerate(transactions, 2):
            low_conf   = tx.get('_low_conf')
            bal_flag   = tx.get('_balance_flag')
            auto_fixed = bool(bal_flag and bal_flag.startswith('auto-corrected'))
            flag_text  = bal_flag or ('Low OCR confidence' if low_conf else '')

            if auto_fixed:
                fill = fix_fill
            elif low_conf or bal_flag:
                fill = warn_fill
            else:
                fill = alt_fill if row_idx % 2 == 0 else None

            values = self.to_excel_row(tx) + [flag_text]
            for col, (val, aln) in enumerate(zip(values, aligns), 1):
                c = ws.cell(row=row_idx, column=col, value=val)
                c.font, c.alignment, c.border = body_font, aln, border
                if fill:
                    c.fill = fill
                if col in num_cols and val is not None:
                    c.number_format = '#,##0.00'
            ws.row_dimensions[row_idx].height = 30

        for col_letter, width in self.EXCEL_COL_WIDTHS.items():
            ws.column_dimensions[col_letter].width = width
        flag_col_letter = chr(ord('A') + n_cols - 1)
        ws.column_dimensions[flag_col_letter].width = 24
        ws.auto_filter.ref = f"A1:{flag_col_letter}{len(transactions)+1}"

        # ── Summary sheet ───────────────────────────────────────────────────
        ss = wb.create_sheet('Summary')
        rows = self.excel_summary_rows(transactions, meta)
        for r, (label, value) in enumerate(rows, 1):
            lc = ss.cell(row=r, column=1, value=label)
            vc = ss.cell(row=r, column=2, value=value)
            lc.font = Font(name='Arial', bold=True, size=10)
            vc.font = Font(name='Arial', size=10)
            if isinstance(value, float):
                vc.number_format = '#,##0.00'
        ss.column_dimensions['A'].width = 22
        ss.column_dimensions['B'].width = 35

        wb.save(output_path)
        return output_path
